from selenium import webdriver
import time
from tkinter import messagebox
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import csv
import os
import pyautogui
from openpyxl import load_workbook
import shutil
import win32gui
from tkinter import messagebox
w=win32gui
path = "C:\\GTP"
isExist = os.path.exists(path)
if not isExist:
   os.makedirs(path)
path = "C:\\GTP\\Output"
isExist = os.path.exists(path)
if not isExist:
   os.makedirs(path)
path = "C:\\GTP\\Output\\Processed"
isExist = os.path.exists(path)
if not isExist:
   os.makedirs(path)
else:
    shutil.rmtree(path)
    os.makedirs(path)
path = "C:\\GTP\\Output\\Unprocessed"
isExist = os.path.exists(path)
if not isExist:
   os.makedirs(path)
else:
    shutil.rmtree(path)
    os.makedirs(path)
if os.path.exists("C:\\GTP\\Output\\Processed.zip"):
  os.remove("C:\\GTP\\Output\\Processed.zip")
if os.path.exists("C:\\GTP\\Output\\Unprocessed.zip"):
  os.remove("C:\\GTP\\Output\\Unprocessed.zip")
email_body = ''
TO = ''
FROM = ''
filename = r"C:\GTP\Config.xlsx"
wb =load_workbook(filename)
ws = wb.active
vr = 0
message = ''
scount = '0'
email_message = ''
for row in ws.iter_rows(2, ws.max_row):
    for cell in row:
        if cell.value == 'EmailTo':
            cell1 = cell.row
            TO = ws.cell(row = cell1, column = 2).value
            print(TO)
        if cell.value == 'EmailFrom':
            cell1 = cell.row
            FROM = ws.cell(row = cell1, column = 2).value
            print(FROM)
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
for row in ws.iter_rows(2, ws.max_row):
    for cell in row:
        if cell.value == 'URLS':
            cell1 = cell.row
            print(cell1)
            scount = int(scount) + int(1)
            link = ws.cell(row = cell1, column = 2).value
            print(link)
            username = ws.cell(row = cell1, column = 3).value
            print(username)
            password = ws.cell(row = cell1, column = 4).value
            print(password)
            wintitle = ws.cell(row = cell1, column = 5).value
            print(wintitle)
            #try:
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
            driver.get(link)
            driver.maximize_window()
            time.sleep(10)
            message = message + '\n' + '------------------' + '\n' + wintitle + '\n' + '------------------' + '\n'
            if w.GetWindowText(w.GetForegroundWindow()) == wintitle:
               driver.find_element("xpath", "//input[@id='txtUsrName']").send_keys(username)
                  #pyautogui.write(idnumber, interval=0.25)
                  #time.sleep(2)
                  #pyautogui.press('enter')
               driver.find_element("xpath", "//input[@id='txtPassword']").send_keys(password)
               driver.find_element("xpath", "//*[@id='loginTable']/tbody/tr[4]/td/input").click()
               time.sleep(10)
               if w.GetWindowText(w.GetForegroundWindow()) == 'Search - Google Chrome':
                  driver.find_element("xpath", "//*[@id='SearchCriteria_chzn']/a/span").click()
                  time.sleep(2)
                  driver.find_element("xpath", "//*[@id='SearchCriteria_chzn_o_3']").click()
                  driver.find_element("xpath", "//*[@id='SearchText']").send_keys('d')
                  driver.find_element("xpath", "//*[@id='submitSearch']").click()
                  driver.find_element("xpath", "//*[@id='searchGrid']/tbody/tr[1]").click()
                  time.sleep(4)
                  #idnumber = driver.find_element("xpath", "//*[@id='contractDetails']/table/tbody/tr[2]/td[2]").value('innerHTML')
                  idnumber = driver.find_element(by = By.XPATH, value = '//*[@id="contractDetails"]/table/tbody/tr[2]/td[2]')                  
                  idnumber = idnumber.get_attribute('innerHTML')
                  idnumber = idnumber.strip()
                  print('id number =',idnumber)                  
                  find = driver.find_element("xpath", "//*[@id='contractDetails']/table/tbody/tr[2]/td[1]").is_enabled()
                  if find == True:
                     message = message + 'Contract Details for ' + idnumber + '- Working' + '\n'
                  else:
                     message = message + 'Contract Details for ' + idnumber + '- Not Working' + '\n'
                  driver.find_element("xpath", "//*[@id='tradingIdv']").click()
                  time.sleep(1)
                  driver.find_element("xpath", "//div[@id='tradingIdvSubmenu']/a[1]").click()
                  time.sleep(4)
                  find = driver.find_element("xpath", "//*[@id='contractDiv']/div[1]/span[1]").is_enabled()
                  if find == True:
                     message = message + 'Transfers Details for ' + idnumber + '- Working' + '\n'
                  else:
                     message = message + 'Transfers Details for ' + idnumber + '- Not Working' + '\n'
                  driver.find_element("xpath", "//*[@id='tradingIdv']").click()
                  time.sleep(1)
                  driver.find_element("xpath", "//*[@id='tradingIdvSubmenu']/a[2]").click()
                  time.sleep(4)
                  find = driver.find_element("xpath", "//*[@id='contractReallocationDiv']/div[1]/span[1]").is_enabled()
                  if find == True:
                     message = message + 'REALLOCATION Details for ' + idnumber + '- Working' + '\n'
                  else:
                     message = message + 'REALLOCATION Details for ' + idnumber + '- Not Working' + '\n'
                  driver.find_element("xpath", "//*[@id='tradingIdv']").click()
                  time.sleep(2)
                  driver.find_element("xpath", "//*[@id='tradingIdvSubmenu']/a[3]").click()
                  time.sleep(4)
                  find = driver.find_element("xpath", "//*[@id='fundReallocationDiv']/div[1]/span").is_enabled()
                  if find == True:
                     message = message + 'FUTURE ALLOCATION Details for ' + idnumber + '- Working' + '\n'
                  else:
                     message = message + 'FUTURE ALLOCATION Details for ' + idnumber + '- Not Working' + '\n'
                  driver.find_element("xpath", "//*[@id='reporting']").click()
                  time.sleep(1)
                  driver.find_element("xpath", "//*[@id='reportingSubmenu']/a[1]").click()
                  time.sleep(4)
                  find = driver.find_element("xpath", "//*[@id='mainwrapper']/div[2]").is_enabled()
                  if find == True:
                     message = message + 'PENDING TRANSACTION Details for ' + idnumber + '- Working' + '\n'
                  else:
                     message = message + 'PENDING TRANSACTION Details for ' + idnumber + '- Not Working' + '\n'
               else:
                  message = message + link + ' - ' + 'Unable to login to the site'
            else:
               message = message + link + ' - ' + 'Site is Down'
            #except:
            #message = 'Error occur need to check manually'
            
            email_message = email_message+' - '+message
            email_body = email_body+'\n'+email_message
            #print(email_body)
            print(cell1)
            path = 'C:\GTP\Output\Processed' + '\\' + str(scount) + '.png'
            print(path)
            driver.save_screenshot(path)
            driver.close()
            #print(message)
            #messagebox.showinfo(message)
        elif cell.value == 'EmailFrom':
            cell1 = cell.row
            TO = ws.cell(row = cell1, column = 2).value
            print(TO)
        elif cell.value == 'EmailTo':
            cell1 = cell.row
            FROM = ws.cell(row = cell1, column = 2).value
            print(FROM)
'''SERVER = "outlook.sbl.com"
TO = ['nagender11.yadav@nttdata.com', 'nagender.yadav@se2.com']
SUBJECT = "Fast - Sunday Health Check"
TEXT = 'Hi Team,\n\nPlease find the Fast - Health check results.\n\nLink Check -\n'+(email_body)+'\n\nRegards,\nRPA Team'

message = """From: %s\r\nTo: %s\r\nSubject: %s\r\n\

%s
""" % (FROM, ", ".join(TO), SUBJECT, TEXT)
import smtplib
server = smtplib.SMTP(SERVER,25)
server.sendmail(FROM, TO, message)
server.quit()'''
print(message)
shutil.make_archive('C:\\GTP\\Output\\Processed.zip', 'zip', 'C:\\GTP\\Output\\Processed')
shutil.make_archive('C:\\GTP\\Output\\Unprocessed.zip', 'zip', 'C:\\GTP\\Output\\Unprocessed')
import win32com.client as win32
outlook = win32.Dispatch('outlook.application')
mail=outlook.CreateItem(0)
mail.To='nagender11.yadav@nttdata.com;nagender.yadav@se2.com'
mail.Subject='GTP - Sunday Health Check'
mail.HTMLBody='Hi Team,\n\nPlease find the GTP - Health check results.\n\nLink Check -\n'+ message +'\n\nRegards,\nRPA Team' #this field is optional
mail.Body='Hi Team,\n\nPlease find the GTP - Health check results.\n\nLink Check -\n'+ message +'\n\nRegards,\nRPA Team'
# To attach a file to the email (optional):
#if os.path.exists("C:\\Fast\\Output\\Processed.zip"):
if os.path.exists("C:\GTP\Output\Processed.zip"):
    attachment="C:\GTP\Output\Processed.zip"
    mail.Attachments.Add(attachment)
'''if os.path.exists("C:\\Fast\\Output\\Unprocessed.zip"):
   attachment="C:\\Fast\\Output\\Unprocessed.zip"
   mail.Attachments.Add(attachment)'''
mail.Send()
print('sent')
